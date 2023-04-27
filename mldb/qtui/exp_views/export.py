from collections import defaultdict
import os.path

from PySide6.QtWidgets import QPushButton, QVBoxLayout, QFileDialog, QMessageBox

import psycopg2.sql as sql
import numpy as np
from openpyxl import Workbook

from .view_base import BaseExpView
from ..db_iop import DBQuery, DBExpDetails, DBExpQualResults


class ExportData(BaseExpView):
    def __init__(self, *expids):
        super().__init__(*expids)
        self.layout = QVBoxLayout(self)
        self.btn = QPushButton("Export")
        self.btn.setEnabled(False)
        self.btn.clicked.connect(self.export_data)
        self.layout.addWidget(self.btn)

        self.last_dir = "/Users/christopherboyle/gits/mdpc/CLD2QC/exported_exp_data"
        self.data = {}
        self.exp_groups = {}
        self.workbook = Workbook()
        self.metrics_sheet = self.workbook.active
        self.metrics_sheet.title = 'Metrics'
        self.expids_to_get = set(self.expids)
        self.current_expid = None

        self.refresh()

    def refresh(self):
        self.btn.setEnabled(False)
        self.expids_to_get = set(self.expids) if len(self.expids) < 10 else {}
        self.get_expdetails_and_qualres_then_metrics()
    
    def get_metrics(self):
        self.data = {}
        q = sql.SQL("SELECT * FROM METRICS WHERE {condition};")
        condition = sql.SQL("")
        for i, expid in enumerate(self.expids):
            expid = sql.Literal(expid)
            if i:
                condition += sql.SQL(" OR ")
            condition += sql.SQL(" EXPID={e} ").format(e=expid)
            # fewer rows return, but is slower!
            # condition += sql.SQL(
            #     "((EXPID, EPOCH) IN (SELECT EXPID, max(EPOCH) FROM METRICS WHERE EXPID={e} GROUP BY EXPID))"
            # ).format(e=expid)
        DBQuery(q.format(condition=condition), self.metrics_returned).start()
    
    def get_expdetails_and_qualres_then_metrics(self):
        if self.expids_to_get:
            self.current_expid = self.expids_to_get.pop()
            DBExpQualResults(self.current_expid, self.qualres_returned).start()
    
    def qualres_returned(self, plotids_and_data):
        for plotid, data in plotids_and_data:
            sheet = self.workbook.create_sheet('...' + self.current_expid[-20:] + ' qualres')
            if len(np.array(data["data"][0]["output"]).flatten()) > 1:
                data = data['data']
                nplots = len(np.array(data[0]["output"]).flatten())
                try:
                    subids, tag = plotid.split(";")
                    subids = subids.split("-")
                    subids = [(pid + ";" + tag) for pid in subids]
                    assert len(subids) == nplots
                except (IndexError, AssertionError, ValueError):
                    subids = ['conc', 'd10', 'd50', 'd90']  # HARDCODED!! TODO!
                assert len(subids) == nplots
                
                epoch = max([d["epoch"] for d in data])
                data = [d for d in data if d["epoch"] == epoch]

                for i, subid in enumerate(subids):
                    sheet.cell(row=1, column=i*3 + 1).value = f'Target ({subid})'
                    sheet.cell(row=1, column=i*3 + 2).value = f'Predicted ({subid})'
                    targets = [d["target"][0][i] for d in data]
                    outputs = [d["output"][0][i] for d in data]
                    for j, (t, o) in enumerate(zip(targets, outputs)):
                        sheet.cell(row=j+2, column=i*3+1).value = t
                        sheet.cell(row=j+2, column=i*3+2).value = o
            else:
                print('todo')
        DBExpDetails(self.current_expid, self.expdetails_returned).start()
    
    def expdetails_returned(self, expdeets):
        if 'losses' in expdeets:
            sheet = self.workbook.create_sheet('...' + expdeets['expid'][-20:] + ' losses')

            sheet.cell(row=1, column=1).value = 'Exp. ID:'
            sheet.cell(row=1, column=2).value = expdeets['expid']
            sheet.cell(row=1, column=3).value = 'Status:'
            sheet.cell(row=1, column=4).value = expdeets['status']

            sheet.cell(row=3, column=1).value = 'Epochs'
            sheet.cell(row=3, column=2).value = 'Training Loss'

            te = expdeets['losses']['train']['epoch']
            tv = expdeets['losses']['train']['loss']
            if 'valid' in expdeets['losses']:
                sheet.cell(row=3, column=3).value = 'Valid Loss'
                ve = expdeets['losses']['valid']['epoch']
                vv = expdeets['losses']['valid']['loss']
            else:
                ve = vv = None
            
            if 'lrs' in expdeets:
                sheet.cell(row=3, column=4).value = 'Learning Rate'
                le = expdeets['lrs']['epochs']
                lv = expdeets['lrs']['lrs']
            else:
                le = lv = None
            
            epochs = te if ve is None else sorted(set([*te, *ve]))

            for ei, e in enumerate(epochs):
                sheet.cell(row=ei+4, column=1).value = e
                if e == 0:
                    sheet.cell(row=ei+4, column=6).value = 'i.e., before training has commenced.'
                try:
                    ti = te.index(e)
                    tvi = tv[ti]
                    sheet.cell(row=ei+4, column=2).value = tvi
                except ValueError:
                    pass
                if vv is not None:
                    try:
                        vi = ve.index(e)
                        vvi = vv[vi]
                        sheet.cell(row=ei+4, column=3).value = vvi
                    except:
                        pass
                if lv is not None:
                    try:
                        li = le.index(e)
                        lvi = lv[li]
                        sheet.cell(row=ei+4, column=4).value = lvi
                    except:
                        pass
        
        if self.expids_to_get:
            self.get_expdetails_and_qualres_then_metrics()
        else:
            self.get_metrics()

    def metrics_returned(self, rows):
        data_by_exp_epoch = defaultdict(lambda: defaultdict(dict))
        for expid, epoch, kind, value in rows:
            data_by_exp_epoch[expid][epoch][kind] = value

        latest_data_by_exp = {}
        for expid, epoch_data in data_by_exp_epoch.items():
            latest_data = epoch_data[sorted(epoch_data)[-1]]
            latest_data_by_exp[expid] = latest_data

        self.data = latest_data_by_exp

        self.get_groups()

    def get_groups(self):
        q = sql.SQL("SELECT * FROM EXPGROUPS WHERE {condition};")
        condition = sql.SQL("")
        for i, expid in enumerate(self.expids):
            expid = sql.Literal(expid)
            if i:
                condition += sql.SQL(" OR ")
            condition += sql.SQL(" EXPID={e} ").format(e=expid)
        DBQuery(q.format(condition=condition), self.groups_returned).start()

    def groups_returned(self, rows):
        primary_groups_by_exp = {}
        for expid, group in rows:
            if "=" in group:
                primary_groups_by_exp[expid] = group

        for expid, pgroup in primary_groups_by_exp.items():
            gdata = {}
            for item in pgroup.split(";"):
                kv = item.split("=")
                if len(kv) > 1:
                    k, v = kv[0], '='.join(kv[1:])
                    if isinstance(v, str):
                        v = f'"{v}"'
                    gdata[k] = v
            if expid in self.data:
                mdata = self.data[expid]
                self.data[expid] = {**gdata, **mdata}

        self.btn.setEnabled(True)

    def export_data(self):
        fn, _ = QFileDialog.getSaveFileName(
            self, "location to save data", self.last_dir, "*.xlsx"
        )
        if fn:
            self.last_dir = os.path.dirname(fn)
            self.save_to_file(fn)

    def data_to_sheet(self):
        rows = sorted(self.data)
        columns = list()  # list used instead of set to preserve order
        for edata in self.data.values():
            for k in edata:
                if k not in columns:
                    columns.append(k)
        
        for i, column in enumerate(columns):
            self.metrics_sheet.cell(row=1, column=i+2).value = column

        table = []
        for r, row in enumerate(rows):
            row_data = [row]
            for c, col in enumerate(columns):
                try:
                    row_data.append(self.data[row][col])
                except KeyError:
                    row_data.append("")
            table.append(row_data)
        
        for r, row in enumerate(table):
            for c, cell in enumerate(row):
                self.metrics_sheet.cell(row=r+2, column=c+1).value = cell

    def save_to_file(self, fn):
        self.data_to_sheet()
        self.workbook.save(fn)

        mbox = QMessageBox(self)
        mbox.setWindowTitle("Done!")
        mbox.setText(f"Data has been written to {fn}")
        mbox.show()
